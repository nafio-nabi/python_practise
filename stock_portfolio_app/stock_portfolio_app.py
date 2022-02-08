import openpyxl

# Open file
workbook = openpyxl.load_workbook("stock_portfolio.xlsx")
portfolio = workbook["portfolio_one"]

# Calculate total value for each stock in portfolio.
for row in range(2, portfolio.max_row + 1):
    stock_price = portfolio.cell(row, 4).value
    stock_quantity = portfolio.cell(row, 5).value
    stock_value = stock_price * stock_quantity
    
    # Write total value for each stock in portfolio to xlsx file.
    value_row = portfolio.cell(row, 6).value
    value_row = stock_value

# Save file
workbook.save("stock_portfolio_updated.xlsx")